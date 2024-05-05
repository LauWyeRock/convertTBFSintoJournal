import pandas as pd
from openpyxl import load_workbook
import streamlit as st
from functools import reduce
import tempfile
import shutil



# def process_file(file_path, sheet_name, skip_rows, last_row_to_skip, column_names=None):

#     with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
#         shutil.copyfileobj(file_path, temp_file)
#         temp_file_path = temp_file.name

#     excel_app = xw.App(visible=False)
#     excel_book = None
#     try:
#         excel_book = excel_app.books.open(temp_file_path)
#         excel_book.save()
#     finally:
#         excel_book.close()
#         excel_app.quit()

#     wb = load_workbook(temp_file_path, data_only=True)
#     ws = wb[sheet_name]

#     rows = list(ws.iter_rows(values_only=True, min_row=skip_rows + 1, max_row=ws.max_row - last_row_to_skip))
#     data = pd.DataFrame(rows, columns=column_names if column_names else None)

#     if not column_names:
#         header_row = data.iloc[0]  
#         header_row[0] = 'Account' 
#         data.columns = header_row
#         data = data.iloc[1:]

#     data.reset_index(drop=True, inplace=True)
#     for col in data.columns[1:]:  
#         data[col] = pd.to_numeric(data[col], errors='coerce')
#     data.fillna(0, inplace=True)
#     data = data[(data['Account'] != "0")]

#     return data

def process_file(file_path, sheet_name, skip_rows, last_row_to_skip, column_names=None):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        shutil.copyfileobj(file_path, temp_file)
        temp_file_path = temp_file.name

    if column_names is None:
        data = pd.read_excel(temp_file_path, sheet_name=sheet_name, 
                             skiprows=skip_rows, 
                             skipfooter=last_row_to_skip, 
                             header=None)
        data.columns = data.iloc[0]
        data = data[1:]
    else:
        data = pd.read_excel(temp_file_path, sheet_name=sheet_name, 
                             skiprows=skip_rows, 
                             skipfooter=last_row_to_skip, 
                             header=None, 
                             names=column_names)

    data.reset_index(drop=True, inplace=True)
    if 'Account' in data.columns: 
        data['Account'] = data['Account'].astype(str).str.strip()
        data = data[(data['Account'] != "0")]
    for col in data.columns[1:]:  
        data[col] = pd.to_numeric(data[col], errors='coerce')

    return data

def main():
    st.title("Financial Data Processor")

    tb_file = st.file_uploader("Upload Trial Balance", type=['xlsx'])
    bs_file = st.file_uploader("Upload Balance Sheet", type=['xlsx'])
    pl_file = st.file_uploader("Upload Profit & Loss", type=['xlsx'])

    if st.button("Process Files"):
        if tb_file and bs_file and pl_file:
            tb_data = process_file(tb_file, 'Trial Balance', 4, 1, ['Account', 'TB Debit', 'TB Credit'])
            bs_data = process_file(bs_file, 'Balance Sheet', 4, 1, ['Account', 'BS Amount'])
            pl_data = process_file(pl_file, 'Profit and Loss', 4, 1, ['Account', 'P&L Amount'])


            dfs = [tb_data.set_index('Account'), bs_data.set_index('Account'), pl_data.set_index('Account')]
            combined_data = reduce(lambda left, right: pd.merge(left, right, on='Account', how='outer'), dfs).reset_index()
            combined_data['Account'] = combined_data['Account'].str.strip()

            combined_data['Non_Null_Count'] = combined_data.count(axis=1)
            combined_data = combined_data.sort_values(by='Non_Null_Count', ascending=False)
            combined_data = combined_data.drop_duplicates(subset='Account', keep='first').reset_index(drop=True)
            combined_data = combined_data.drop(columns='Non_Null_Count')

            combined_data = combined_data[(combined_data['Account'].astype(str) != '0') & (combined_data['Account'].notna()) & (combined_data['Account'].astype(str) != 'nan')]
            combined_data = combined_data[~((combined_data[['TB Debit', 'TB Credit', 'BS Amount', 'P&L Amount']].isna() | (combined_data[['TB Debit', 'TB Credit', 'BS Amount', 'P&L Amount']] == 0)).sum(axis=1) == 4)]

            combined_data[['Account Code', 'Account Name']] = combined_data['Account'].str.extract(r'(?:(\d+(?:\.\d+)*)\s+)?(.*)') 

            output_df = pd.DataFrame({
                'Journal Reference': 'FYE2023 Conversion: Transfer Balance',
                'Contact': None,
                'Date': None,
                'Account': combined_data['Account Name'],
                'Description': None,
                'Tax Included in Amount': None,
                'Debit Amount': combined_data['TB Debit'],
                'Credit Amount': combined_data['TB Credit'],
                'raw_account': combined_data['Account'],
                'raw_TB-dedit': combined_data['TB Debit'],
                'raw_TB-credit': combined_data['TB Credit'],
                'raw_BS-amount': combined_data['BS Amount'],
                'raw_P&L-amount': combined_data['P&L Amount']
            })

            st.dataframe(output_df)
            csv = output_df.to_csv(index=False)
            st.download_button("Download CSV", csv, "financial_data.csv", "text/csv", key='download-csv')
        else:
            st.error("Please upload all files to process.")

if __name__ == "__main__":
    main()
